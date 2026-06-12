"""Export stage: write processed records to a ready-to-use Excel file.

The end of the pipeline. It takes the records, turns them into a DataFrame (via
the shared `records_to_frame`), and writes a `.xlsx` that someone on the desk can
open and use immediately: a bold frozen header row, auto-sized columns, an
auto-filter, and a small "_meta" sheet recording where/when the data came from.

Excel is written with openpyxl through pandas' ExcelWriter.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import SourceConfig
from .records import ScrapeRecord, records_to_frame

# Cap auto-sizing so one long URL cell can't make a column absurdly wide.
_MAX_COL_WIDTH = 60


def _autosize_and_style(worksheet) -> None:
    """Bold + freeze the header, add an auto-filter, and fit column widths."""
    for cell in worksheet[1]:  # worksheet[1] = row 1 (openpyxl rows are 1-based!)
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"  # freeze everything ABOVE A2 -> header stays on scroll
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        # build an excel range like "A1:C30". get_column_letter turns a column NUMBER
        # into its letter (1->A, 2->B, 27->AA). auto_filter adds the dropdown arrows.
        last_col = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{last_col}{worksheet.max_row}"

    for column_cells in worksheet.columns:  # iterate one column (all its cells) at a time
        # longest text in the column = how wide to make it. generator skips empty cells;
        # default=0 protects max() from erroring on a fully-empty column.
        longest = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        letter = get_column_letter(column_cells[0].column)
        # +2 for a little padding; min(..., cap) stops one long URL blowing it out.
        worksheet.column_dimensions[letter].width = min(longest + 2, _MAX_COL_WIDTH)


def export(
    records: list[ScrapeRecord],
    config: SourceConfig,
    output: str | None = None,
) -> Path:
    """Write `records` to an Excel file; return the path written.

    `output` overrides `config.output` when given (e.g. from the CLI).
    """
    # `output or config.output` = use the CLI override if given, else the config's path.
    # expanduser() turns a leading "~" into the real home dir.
    path = Path(output or config.output).expanduser()
    if path.parent != Path():  # Path() is "."; skip mkdir if output is in the cwd
        # parents=True makes intermediate folders; exist_ok=True = don't error if it's there.
        path.parent.mkdir(parents=True, exist_ok=True)

    frame = _excel_safe(records_to_frame(records))

    # `with ... as writer`: the context manager SAVES + closes the file on exit. if you
    # forget the `with`, the .xlsx may never actually get written to disk.
    with pd_writer(path) as writer:
        sheet_name = (config.name or "data")[:31]  # Excel caps sheet names at 31 chars
        frame.to_excel(writer, sheet_name=sheet_name, index=False)  # index=False: no 0,1,2 col
        _autosize_and_style(writer.sheets[sheet_name])  # the sheet object pandas just made

        meta = _meta_frame(records, config)
        meta.to_excel(writer, sheet_name="_meta", index=False)  # second sheet in same file
        _autosize_and_style(writer.sheets["_meta"])

    return path


def _excel_safe(frame):
    """Strip timezones from datetime columns -- Excel rejects tz-aware values.

    `scraped_at` is recorded in UTC; we drop the tzinfo (keeping the UTC clock
    time) so openpyxl will write it.
    """
    import pandas as pd

    for col in frame.columns:
        series = frame[col]  # one column of the table (a pandas Series)
        # case 1: a proper tz-aware datetime column -> .dt.tz_localize(None) drops the tz.
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            frame[col] = series.dt.tz_localize(None)
        # case 2: an "object" column (mixed python values) might hold tz-aware datetimes;
        # .map applies a function to each cell, stripping tzinfo only where present.
        elif series.dtype == object:
            frame[col] = series.map(
                lambda v: v.replace(tzinfo=None)
                if isinstance(v, datetime) and v.tzinfo is not None
                else v  # leave non-datetime values untouched
            )
    return frame


def pd_writer(path: Path):
    """An openpyxl-backed pandas ExcelWriter (kept here to localise the import)."""
    import pandas as pd

    # engine="openpyxl" = use the openpyxl library to actually write the .xlsx (it's
    # what lets us style cells afterwards via writer.sheets[...]).
    return pd.ExcelWriter(path, engine="openpyxl")


def _meta_frame(records: list[ScrapeRecord], config: SourceConfig):
    """A tiny provenance table: what was scraped, from where, and when."""
    import pandas as pd

    return pd.DataFrame(
        [
            {"key": "source", "value": config.name},
            {"key": "type", "value": config.type},
            {"key": "url", "value": config.url},
            {"key": "rows", "value": len(records)},
            {"key": "exported_at", "value": datetime.now(timezone.utc).isoformat()},
        ]
    )
